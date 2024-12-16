from flask import Flask, render_template, request, jsonify, send_file
import pandas as pd
import matplotlib.pyplot as plt
import io
import base64
from docx import Document
from docx.shared import Inches
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from google.oauth2.service_account import Credentials
import os.path
import pickle
import json
import csv
from io import StringIO
import openpyxl
from openpyxl import Workbook
from io import BytesIO
from flask import send_file
from dotenv import load_dotenv  # Thêm dòng này
app = Flask(__name__)

# Tắt HTTPS requirement cho OAuth 2.0
os.environ['OAUTHLIB_INSECURE_TRANSPORT'] = '1'


# Biến toàn cục để lưu DataFrame
df = None
load_dotenv('google_sheet.env')


# Phạm vi quyền truy cập Google Sheets
SCOPES = [
    'https://www.googleapis.com/auth/spreadsheets.readonly',
    'https://www.googleapis.com/auth/userinfo.email',
    'https://www.googleapis.com/auth/userinfo.profile',
    'openid',  # Thêm 'openid' nếu cần
]

# ID của Google Sheet (thay thế bằng ID của bạn)
SPREADSHEET_ID = os.getenv('SPREADSHEET_ID')
RANGE_NAME = os.getenv('RANGE_NAME')

def get_google_sheets_credentials():
    """Lấy và cập nhật credentials từ nội dung JSON trong biến môi trường."""
    # creds = None
    # if os.path.exists('token.pickle'):
    #     with open('token.pickle', 'rb') as token:
    #         creds = pickle.load(token)
    # if not creds or not creds.valid:
    #     if creds and creds.expired and creds.refresh_token:
    #         creds.refresh(Request())
    #     else:
    #         # Lấy nội dung JSON từ biến môi trường
    #         credentials_info = json.loads(os.getenv("GOOGLE_CREDENTIALS"))
    #         flow = InstalledAppFlow.from_client_config(credentials_info, SCOPES)
    #         creds = flow.run_local_server(port=5000)
    #         with open('token.pickle', 'wb') as token:
    #             pickle.dump(creds, token)
    # return creds
    """Sử dụng Service Account từ biến môi trường."""
    try:
        # Lấy nội dung JSON từ biến môi trường
        credentials_info = json.loads(os.getenv("GOOGLE_SERVICE_ACCOUNT"))
        
        # Xác thực với Service Account
        creds = Credentials.from_service_account_info(
            credentials_info,
            scopes=SCOPES
        )
        return creds
    except Exception as e:
        raise Exception(f"Error setting up Service Account: {e}")

def fetch_data_from_sheets():
    global df
    try:
        creds = get_google_sheets_credentials()
        service = build('sheets', 'v4', credentials=creds)

        sheet = service.spreadsheets()
        result = sheet.values().get(
            spreadsheetId=SPREADSHEET_ID,
            range=RANGE_NAME
        ).execute()

        values = result.get('values', [])
        if not values:
            raise Exception('No data found in the sheet.')

        headers = values[0]
        data = values[1:]
        df = pd.DataFrame(data, columns=headers)


        return df
    except Exception as e:
        print(f"Error fetching data: {e}")
        raise Exception(f'Error fetching data from Google Sheets: {str(e)}')

@app.route('/')
def home():
    return render_template('index.html', table=None, plot_url=None)

@app.route('/data')
def show_data():
    try:
        df = fetch_data_from_sheets()
        table_html = df.to_html(index=False, classes='table table-striped table-bordered', justify='center')
        return render_template('upload.html', table=table_html, plot_url=None)
    except Exception as e:
        return f'Error: {str(e)}'

# Route vẽ biểu đồ
@app.route('/plot', methods=['POST'])
def plot():
    global df
    if df is None:
        return jsonify({'error': 'No data to plot'}), 400

    try:
        # Tính số lượng tồn tại của từng sản phẩm
        product_counts = df['MẶT HÀNG'].value_counts()

        # Tạo đồ thị bar
        img = io.BytesIO()
        plt.figure(figsize=(10, 6))
        plt.bar(product_counts.index, product_counts.values, color='skyblue') 
        plt.xlabel('PRODUCTS NAME')
        plt.ylabel('QUANTITY')
        plt.title('SỐ LƯỢNG HÀNG HÓA')
        plt.tight_layout()
        plt.savefig(img, format='png')
        img.seek(0)
        plot_url = base64.b64encode(img.getvalue()).decode('utf8')

        return jsonify({'plot_url': plot_url})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/plot_by_day', methods=['POST'])
def plot_by_day():
    global df
    if df is None:
        return jsonify({'error': 'No data to plot'}), 400

    try:
        # Chuyển cột TIME sang định dạng datetime và trích xuất ngày
        df['DATE'] = pd.to_datetime(df['TIME']).dt.date

        # Lọc dữ liệu theo từng ngày
        unique_dates = df['DATE'].unique()
        plot_urls = {}

        for date in unique_dates:
            # Lọc dữ liệu của ngày hiện tại
            daily_data = df[df['DATE'] == date]

            # Tính số lượng sản phẩm
            product_counts = daily_data['MẶT HÀNG'].value_counts()

            # Tạo biểu đồ
            img = io.BytesIO()
            plt.figure(figsize=(10, 6))
            plt.bar(product_counts.index, product_counts.values, color='skyblue')
            plt.xlabel('PRODUCTS NAME')
            plt.ylabel('QUANTITY')
            plt.title(f'SỐ LƯỢNG HÀNG HÓA - {date}')
            plt.tight_layout()
            plt.savefig(img, format='png')
            img.seek(0)

            # Chuyển hình ảnh thành base64
            plot_urls[str(date)] = base64.b64encode(img.getvalue()).decode('utf8')

        return jsonify({'plot_urls': plot_urls})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
 
# Route để xuất toàn bộ nội dung ra file CSV
@app.route('/export', methods=['POST'])
def export_to_csv():
    global df
    if df is None:
        return jsonify({'error': 'No data available to export'}), 400

    try:
        # Lưu dữ liệu DataFrame thành file CSV trong bộ nhớ, sử dụng mã hóa UTF-8 với BOM
        csv_io = io.StringIO()
        df.to_csv(csv_io, index=False, encoding='utf-8-sig')  # Sử dụng utf-8-sig để hỗ trợ Unicode
        csv_io.seek(0)

        # Trả về file CSV cho người dùng
        return send_file(
            io.BytesIO(csv_io.getvalue().encode('utf-8-sig')),
            as_attachment=True,
            download_name='product_data.csv',
            mimetype='text/csv'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/update_data', methods=['GET'])
def update_data():
    try:
        df = fetch_data_from_sheets()
        table_html = df.to_html(index=False, classes='table table-striped table-bordered', justify='center')
        return jsonify({'table': table_html})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Route để xuất hóa đơn cho từng ngày ra file CSV
@app.route('/export_invoice', methods=['POST'])
def export_invoice():
    try:
        # Đọc dữ liệu từ Google Sheets
        data = fetch_data_from_sheets()

        # Kiểm tra dữ liệu đầu vào
        if data.empty or not {'TIME', 'QRCODE_DATA', 'GIÁ TIỀN', 'MẶT HÀNG'}.issubset(data.columns):
            return jsonify({"error": "Invalid or missing data from Google Sheets."}), 400

        # Xử lý dữ liệu theo ngày
        invoices = {}
        for _, row in data.iterrows():
            date = row["TIME"].split(" ")[0]  # Lấy ngày
            qrcode = row["QRCODE_DATA"]
            # Xử lý giá tiền
            price = 0
            try:
                price = int(row["GIÁ TIỀN"])
            except (ValueError, TypeError):
                price = 0  # Gán giá trị mặc định nếu lỗi

            if date not in invoices:
                invoices[date] = {}
            if qrcode not in invoices[date]:
                invoices[date][qrcode] = {"SL": 0, "GIÁ": price, "TÊNMH": row["MẶT HÀNG"]}
            invoices[date][qrcode]["SL"] += 1
            invoices[date][qrcode]["TỔNG"] = invoices[date][qrcode]["SL"] * price

        # Tạo file Excel với nhiều sheet
        workbook = Workbook()
        for date, items in invoices.items():
            sheet = workbook.create_sheet(title=date.replace('/', '-'))
            sheet.append(["STT", "TÊNMH", "SL", "GIÁ", "TỔNG"])
            total = 0
            for i, (qrcode, item) in enumerate(items.items(), start=1):
                sheet.append([i, item["TÊNMH"], item["SL"], item["GIÁ"], item["TỔNG"]])
                total += item["TỔNG"]
            sheet.append(["TỔNG TIỀN:", "", "", "", total])

        # Xóa sheet mặc định của openpyxl nếu không sử dụng
        if 'Sheet' in workbook.sheetnames:
            del workbook['Sheet']

        # Lưu file Excel vào bộ nhớ và trả về
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return send_file(output, as_attachment=True, download_name="invoices.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    except Exception as e:
        print(f"Error exporting invoices: {e}")
        return jsonify({"error": str(e)}), 500


# if __name__ == '__main__':
#     app.run(debug=True, port=5001, use_reloader=False)
    
    
# Chạy ứng dụng trên Vercel
app = app

